import openpyxl

# 打开指定的Excel文件
workbook = openpyxl.load_workbook('按市区分类数据_副本.xlsx')

# 循环遍历工作簿，并创建新的工作簿
for worksheet in workbook.worksheets:
    new_workbook = openpyxl.Workbook()

    # 将当前工作簿的名称作为新工作簿的名称
    new_workbook.active.title = worksheet.title

    # 将当前工作簿的内容复制到新工作簿
    for row in worksheet.iter_rows():
        new_workbook.active.append([cell.value for cell in row])

    # 保存新工作簿
    new_workbook.save("{}.xlsx".format(worksheet.title))